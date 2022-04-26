import sys
import logging
from decimal import Decimal
import random
from heapq import merge
from openpyxl import *

logger = logging.Logger('catch_all')
class Generadorv1:
    def __init__(self):
        #inicializa variables
        #self.archivo_base = 'base.xlsx'
        self.archivo_catalogo = 'catalogo.xlsx'
        self.indice_datos = {}
        self.indice_percepciones = {}
        self.indice_deducciones = {}
        self.totald=0    
        self.totalp=0
        self.totaln=0
        self.datos = ["RFC", "HOMOCLAVE", "CURP", "CODIGO", "APELLIDOP", "APELLIDOM", "NOMBRE",
                      "NCUENTA", "IDPAGO", "NOEMPEADO"]
        self.percepciones = ["SUELDO", 'AGUI_SUELDO', "AGUI_COMPEN", "QUINQUENIO", "PREVISION SOCIAL", "DESPENSA", "AEICS","QUINQUENIO",
                             "APOYO_CYD", "PRODUCTIVIDAD", "TRES", "COMPENSACION", "GUARDERIA", "RETRO_PRIMA", "AJUSTE_PRIMA", "NOTAS", "1ER_TRIM","2DO_TRIM","3ER_TRIM","4TO_TRIM","PROD_ANUAL",
                             "PUNT_MENS", "PUNT_PERF", "PERMISOS", "PUNT_ANUAL", "SUBISR", "DIA_MADRE", "DIA_NINO", "DIA_PADRE", "JORNADAS", "RETRO_MENSUAL", "REYES", "CENA",
                             "DIAS", "MENSUAL", "TRIMESTRAL", "TRIMESTRAL2", "PRIMA", "DEVOLUCION", "MENSUAL2", "PRESTAMO", "ACREDITAMIENTO", 'DIA_EMPLEADO']
        self.deducciones = ["ISPT", "CUO_SIN", "ISSSTE", "SEG_MEDICO", "SEG_RETIRO", "PCP", "PHIP",
                            "AHISA", "POTEN", "AHORRO", "MUTUAL", "OTRO", "CUATRO", 'MUTUALISTA',
                            "CINCO","FAR", "DAÑOS", "PENSION", "ISR", 'DANIOS', 'RETRO_CUOTA', 'UTILES', 'CRUZ_ROJA', 'ADEUDO', 'CONCIERTO']
        try:
            #recupera el valor de los argumentos

            #quincena a procesar
            self.qna = int(sys.argv[1])
            #nomina
            self.archivo_nomina = sys.argv[2]
            #base
            self.archivo_base = sys.argv[3]

            #se carga la base de los datos ya existentes
            self._cargar_base_()
            #se carga el catalogo de percepciones/deducciones
            self._carga_catalogo_()
            #se carga la nomina a procesar
            print("Abriendo a procesar ...")
            wb_nomina = load_workbook(filename=self.archivo_nomina)
            hoja_nomina = wb_nomina.worksheets[0]
            #busca los encabezados de los datos
            for indice in range(1, hoja_nomina.max_column+1):
                try:
                    celda = hoja_nomina.cell(row=1, column=indice)
                    if not celda.value is None:
                        #inicializa la bandera de encontrados
                        encontrado = False
                        key = str.upper(celda.value)
                        #far sntea
                        if key == "ISRHOMO":
                            key = "FAR"
                        if key == "BANCO":
                            key = "IDPAGO"
                        if key in self.datos:
                            self.indice_datos[key] = indice-1
                        elif key in self.percepciones:
                            self.indice_percepciones[key] = indice-1
                        elif key in self.deducciones:
                            self.indice_deducciones[key] = indice-1
                        else:
                            print("\t** "+key+" -> no encontrado")
                except Exception as e: 
                    logger.exception('Failed: ' + str(e))
            #inicializa el contador de empleados
            contador_fila = 1
            wb_nomina = Workbook()
            ws_nomina = wb_nomina.active
            #Recorre la nomina a procesar
            for row in hoja_nomina.iter_rows(row_offset=1):
                #inicializa los valores de la fila
                fila = ["2", str(contador_fila).rjust(8, "0"), self.qna, "A", "30FIA0001E", 400]
                #se crea el archivo de excel
                try:
                    if not row[self.indice_datos["RFC"]].value is None:
                        #rfc completo
                        rfc = row[self.indice_datos["RFC"]].value
                        if len(rfc) < 13:
                            rfc= rfc + row[self.indice_datos["HOMOCLAVE"]].value
                        #recupera los valores de la Base
                        if rfc in self.base:
                            base = self.base[rfc]
                        else:
                            print('No se encontro RFC: ' + rfc)
                            pass
                        #inicia la construcción del layout
                        fila.append(rfc)
                        fila.append(row[self.indice_datos["CURP"]].value)
                        fila.append(row[self.indice_datos["NOEMPEADO"]].value)
                        fila.append(base[0])
                        fila.append(base[8])
                        fila.append(base[9])
                        fila.append(base[10])
                        fila.append(base[7])
                        #pagaduria
                        fila.append("3094002900")
                        fila.append(1)
                        ##clave de la plaza
                        clave = [83101, 1003]
                        #clave unidad
                        clave.append(base[1])
                        #clave subunidad
                        clave.append(base[2])
                        clave.append(row[self.indice_datos["CODIGO"]].value)
                        clave.append("00.0")
                        clave.append(base[3])
                        #calcula la clave integrada de la plaza
                        clave_plaza = "".join(map(str, clave))
                        fila.append(clave_plaza)
                        fila += clave
                        fila.append(base[4])
                        fila.append(base[5])
                        fila.append(base[6])

                        vacio = ["", "", "", "", "", ""]
                        ##
                        #se calcula el total de las percepciones
                        ##
                        total_percepciones = Decimal(0.0)
                        conceptos = []
                        contador_conceptos = 0
                        for per in self.indice_percepciones:
                            indice = self.indice_percepciones[per]
                            if not row[indice].value is None:
                                valor = round(Decimal(row[indice].value), 2)
                                if valor > 0:
                                    total_percepciones += valor
                                    contador_conceptos += 1
                                    cat = self.catalogo[per]
                                    conceptos += cat
                                    conceptos += [valor, self.qna, self.qna, "F"]
                                    #conceptos += [valor, '201601', self.qna, "F"]
                                else:
                                    conceptos += vacio

                        ##
                        #se calcula el total de las deducciones
                        ##
                        total_deducciones = Decimal(0.0)
                        for ded in self.indice_deducciones:
                            indice = self.indice_deducciones[ded]
                            if not row[indice].value is None:
                                valor = round(Decimal(row[indice].value), 2)
                                if valor > 0:
                                    total_deducciones += valor
                                    contador_conceptos += 1
                                    if ded == "CUATRO":
                                        if valor == 50:
                                            ded = "MUTUALISTA"
                                        else:
                                            #se define como otros pagos
                                            #porque integra descuento de utiles + mutual
                                            ded = "OTRO"
                                    elif ded == "CINCO":
                                        if valor == 8.5:
                                            #seguro de daños
                                            ded = "DANIOS"
                                        elif valor % 1 == 0:
                                            #pension
                                            ded = "PENSION"
                                            conceptos += vacio
                                        else:
                                            ded = "PENSION"
                                            valor = valor - 8.5
                                            #se agrega seguro de daños
                                            contador_conceptos += 1
                                            cat = self.catalogo["DANIOS"]
                                            conceptos += cat
                                            conceptos += [8.5, self.qna, self.qna, "F"]
                                    #Para todas las deducciones
                                    cat = self.catalogo[ded]
                                    conceptos += cat
                                    conceptos += [valor, self.qna, self.qna, ""]
                                else:
                                    conceptos += vacio
                                    if ded == "CINCO":
                                        conceptos += vacio
                        #se calcula el neto
                        fila.append(total_percepciones-total_deducciones)
                        self.totald += total_deducciones
                        self.totalp += total_percepciones
                        self.totaln += total_percepciones-total_deducciones
                        ## pago
                        pago = row[self.indice_datos["IDPAGO"]].value
                        if not pago == 3:
                            fila.append(row[self.indice_datos["NCUENTA"]].value)
                            fila.append(0)
                        else:
                            fila.append(0)
                            fila.append(0)
                        fila.append('9')
                        #contador
                        fila.append(contador_conceptos)
                        fila += conceptos
                        #print(fila)
                        ws_nomina.append(fila)
                        #se incrementa el contador de filas
                        contador_fila += 1
                except Exception as e:
                    print(e)
            wb_nomina.save("OK/"+str(self.archivo_nomina))
            #imprime totales
            print('total percepciones:')
            print(self.totalp)
            print('total deducciones:')
            print(self.totald)
            print('total neto:')
            print(self.totaln)
        except IndexError:
            print("Uso:")
            print("python v1.py qna nombre_nomina.xlsx")

    def _cargar_base_(self):
        print("se carga la base de los datos ...")
        wbbase = load_workbook(filename=self.archivo_base)
        hojabase = wbbase.worksheets[0]
        self.base = {}
        count = 0
        for row in hojabase.iter_rows(row_offset=2):
            rfc = row[0].value
            if not rfc is None:
                #rfc = rfc[:10]
                self.base[rfc] = (
                    row[1].value,                 #0 nss
                    str(row[6].value).strip(),    #1 unidad
                    str(row[7].value).strip(),    #2 subunidad
                    str(row[9].value).strip(),    #3 folio
                    row[10].value,                #4 nivel_puesto
                    row[11].value,                #5 nivel_sueldo
                    row[12].value,                #6 zona eco
                    row[2].value,                 #7 Completo
                    row[3].value,                 #8 nombre
                    row[4].value,                 #9 paterno
                    row[5].value)                 #10 materno
                count += 1
        print('\t {0} registros -> Ok.'.format(count))

    def _carga_catalogo_(self):
        print("se carga el catalogo de claves ...")
        wbbase = load_workbook(filename=self.archivo_catalogo)
        hojabase = wbbase.worksheets[0]
        self.catalogo = {}
        count = 0
        for row in hojabase.iter_rows(row_offset=2):
            columna = row[0].value
            if not columna is None:
                self.catalogo[columna] = (
                    row[1].value,                 #0 tipo
                    str(row[2].value).strip())    #1 clave
                count += 1
        print('\t {0} registros -> Ok.'.format(count))

    def _carga_nomina_(self):
        pass

#inicializa el metodo principal
generador = Generadorv1()
